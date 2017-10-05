'''
Created on May 14, 2017

@author: Mukadam
'''
from openpyxl import load_workbook
import time,sys
import requests
from urlparse import urlparse
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()
import requests, re, json
from threading import Thread
from Queue import Queue
import time
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()
from pyquery import PyQuery
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()

wb = load_workbook(filename = 'files2.xlsx')

#wb = load_workbook('files.xlsx')
sheet1 = wb.active
url = "https://skrappapi-hello123541.rhcloud.com/api/v2/find"

#querystring = {"firstName":"Chantale","lastName":"Harris","domain":"redseed.com","context":"profile","liv":"nli"}

headers = {
    'x-access-key': "164434961oTzKgbpNNYhwtRAAAAOLoVwIOmu35Tftv",
    'cache-control': "no-cache"
    }
urls = []
for item in range(1,501):
    urls.append(item) 
    
def worker(q):
    while not q.empty():
        item = q.get()
        a =  str(item)+'\t'
        try: 
            
            firstname = sheet1.cell(column=1, row=item).value.split(' ')[0]
            #print firstname
            lastname = sheet1.cell(column=1, row=item).value.split(' ')[1]
            #print lastname
            company_address=  urlparse(sheet1.cell(column=2, row=item).value).netloc.strip('www.')
            #print company_address
            querystring = {"firstName":firstname,"lastName":lastname,"domain":company_address,"context":"profile","liv":"nli"}
    
            response = requests.request("GET", url, headers=headers, params=querystring,verify=False)
            try:
                if response.json()['email_limit_exeeded']:
                    print 'exceeded'
                    sys.exit()
            except Exception,e:
                pass
            a+=(response.json()['email'])
            print a
            
        except Exception,e:
            a+= str(e)
            print a
            #wb.save(files.xlsx')
q = Queue()
map(q.put, urls)
 
startime = time.time()
for i in range(10):
    #print i
    t = Thread(target=worker, args=(q, ))
    t.start()
q.join()
endtime = time.time()
print endtime-startime
