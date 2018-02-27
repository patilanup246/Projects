import urllib
import json, time
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()
from pyquery import PyQuery
from requests_futures.sessions import FuturesSession
session = FuturesSession()
from openpyxl import load_workbook
import time,sys
import requests
from urlparse import urlparse
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()
import requests, re, json
from threading import Thread
from Queue import Queue
import time
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()
from pyquery import PyQuery
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()

base_url = 'http://www.morgansports.com.au/'

# r_main_page  = requests.request('GET',base_url)
# f_main_page_export = open('f_main_page_export.txt','w')

# pq_main_page = PyQuery(r_main_page.text)
# r_main_page.close()
# main_page_all_links = pq_main_page('#top-navigation a')

# for main_page_link in main_page_all_links:
#     print base_url+main_page_link.attrib['href']
#     f_main_page_export.write(base_url+main_page_link.attrib['href'].replace('.html','')+'\n')



# print requests.request('GET','http://www.morgansports.com.au/Boxing-Singlets_c_255-1-4.html',allow_redirects=False).status_code
# f_main_page_export = open('f_main_page_export.txt','r')
# f_second_page_export = open('f_second_page_export.txt','w')
# for link in f_main_page_export.read().split('\n'):
#     i=1
#     while True:
#         print link+'-'+str(i)+'-4.html'
#         r_second_page = requests.request('GET', link+'-'+str(i)+'-4.html',allow_redirects=False)

#         if r_second_page.status_code == 200:
#             pq_second_page = PyQuery(r_second_page.text)
#             all_product_links = pq_second_page('.product-container.middle-item a')
#             r_second_page.close()
#             for product_link in all_product_links:
#                 f_second_page_export.write(base_url+product_link.attrib['href']+'\n')
#                 f_second_page_export.flush()


#         else:
#             break
#         i+=1
# f_second_page_export.close()



excel_headers = ['*ItemCode',
                 'ItemName',
                 'PurchasesDescription',
                 'PurchasesUnitPrice',
                 'PurchasesAccount',
                 'PurchasesTaxRate',
                 'SalesDescription',
                 'SalesUnitPrice',
                 'SalesAccount',
                 'SalesTaxRate',
                 'InventoryAssetAccount',
                 'CostOfGoodsSoldAccount',
                 'Colour',
                 'Size',
                 'URL']

dest_filename = 'morgan.xlsx'
#wb = Workbook()
#ws = wb.active

#ws.title = "Data"

#ws.append(excel_headers)

#wb.save(filename=dest_filename)

# f_second_page_export = open('f_second_page_export.txt', 'r').read().split('\n')
# f_second_page_export = list(set(f_second_page_export))
# print f_second_page_export

product_links_all = []

for i in range(1,4):
    #print i
    resp = requests.get('http://www.morgansports.com.au/product_index.asp?page='+str(i))
    pq = PyQuery(resp.text)
    for link in pq('.product-index-item a'):
        #print link.attrib['href']
        product_links_all.append('http://www.morgansports.com.au/'+link.attrib['href'])



headers = {
        'accept': "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    'accept-encoding': "gzip, deflate, sdch",
    'accept-language': "en-US,en;q=0.8",
    'cache-control': "no-cache",
    'connection': "keep-alive",
    'cookie': "referer=http%3A%2F%2Fwww%2Emorgansports%2Ecom%2Eau%2F; ASPSESSIONIDAQDTRSTB=AIIHFANALMOKMLPFPBJPCMJM; shippingid=94%2B5DS%2FE3nRyJjZqV8R8Gdv3gzCf8HTRVpvYKmEXLBk9Z5gbCpcFShLZF54Nu2lXYLZP2Uav8W4%3D; insurance%5Fselected=; onePageCheckout=1; userid=; lastCat=298; catFilter=; __utmt=1; affiliate=; incompleteorderid=bVNNBgpXUpdgCJj522878; returnUrl=; lastprod=1577; prodbrohist=3%2FQIb0rX216Nlag3FIujCu5PdiIlwems; hasAdvOptions=FeqATj8Kclui2TB%2BGyrkVjMxKehnKBPx0VNZuOvSGxY%3D; thiscat=164; __utma=201642381.567219643.1490924788.1494860042.1494860677.5; __utmb=201642381.34.10.1494860677; __utmc=201642381; __utmz=201642381.1494860677.5.2.utmcsr=google|utmccn=(organic)|utmcmd=organic|utmctr=(not%20provided); __atuvc=0%7C16%2C0%7C17%2C0%7C18%2C0%7C19%2C5%7C20; __atuvs=5919c27cc7e9b350004; 3dvisit=2",
    'host': "www.morgansports.com.au",
    'referer': "http://www.morgansports.com.au/product_index.asp?page=1",
    'upgrade-insecure-requests': "1",
    'user-agent': "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36"
    }
i = 2
file_export_morgan = open('file_export_morgan.txt','w')
def worker(q):
    while not q.empty():
        item = q.get()
        print item
        try:
            r_product_page = requests.request('GET', item, headers=headers, allow_redirects=False)
            pq_product_page = PyQuery(r_product_page.text)
            r_product_page.close()
    
            size = []
            color = []
    
            if (pq_product_page('.container')):
                if (pq_product_page('div.opt-regular:nth-child(1) .label').text().lower() == 'size'):
                    for s in pq_product_page('div.opt-regular:nth-child(1) .radio-option span'):
                        # print s.text
                        size.append(s.text)
                if (pq_product_page('div.opt-regular:nth-child(1) .label').text().lower() == 'colour'):
                    for c in pq_product_page('div.opt-regular:nth-child(1) .radio-option span'):
                        # print c.text
                        color.append(c.text)
                if (pq_product_page('div.opt-regular:nth-child(3) .label').text().lower() == 'size'):
                    for s in pq_product_page('div.opt-regular:nth-child(3) .radio-option span'):
                        # print s.text
                        size.append(s.text)
                if (pq_product_page('div.opt-regular:nth-child(3) .label').text().lower() == 'colour'):
                    for c in pq_product_page('div.opt-regular:nth-child(3) .radio-option span'):
                        # print c.text
                        color.append(c.text)
    
            if len(color) == 0:
                color.append('')
            if len(size) == 0:
                size.append('')
    
            for c in color:
                for s in size:
    #                 ws.cell(row=i, column=1).value = pq_product_page('#product_id').text()
    #                 ws.cell(row=i, column=2).value = pq_product_page('.page_headers').text()
    #                 ws.cell(row=i, column=3).value = pq_product_page('.item[itemprop="description"]').text()
    #                 ws.cell(row=i, column=4).value = pq_product_page('#price').text()
    #                 ws.cell(row=i, column=7).value = pq_product_page('.item[itemprop="description"]').text()
    #                 ws.cell(row=i, column=13).value = c
    #                 ws.cell(row=i, column=14).value = s
    #                 ws.cell(row=i, column=15).value = product_link
    #                 ws.cell(row=i, column=16).value = pq_product_page('.breadcrumbs').text()
    
                    a= pq_product_page('#product_id').text()  +'\t'+ pq_product_page('.page_headers').text().replace('\n', ' ').replace('\r', '') +'\t'+ pq_product_page('.item[itemprop="description"]').text().replace('\n', ' ').replace('\r', '')  +'\t'+  pq_product_page('#price').text()  +'\t'+ c  +'\t'+  s  +'\t'+  item  +'\t'+  pq_product_page('.breadcrumbs').text()+'\n'
                    file_export_morgan.write(a.encode('ascii','ignore'))
                    file_export_morgan.flush()

            #wb.save(filename=dest_filename)
    
        except Exception, e:
            print e
        
        
        
q = Queue()
map(q.put, product_links_all)
 
startime = time.time()
for i in range(20):
    #print i
    t = Thread(target=worker, args=(q, ))
    t.start()
q.join()
endtime = time.time()
print endtime-startime