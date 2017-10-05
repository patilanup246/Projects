from selenium import webdriver
from openpyxl import load_workbook
import time
from selenium.webdriver.common.keys import Keys
wb = load_workbook('Manufacturing Companies_original.xlsx')
ws = wb.active


chromeOptions = webdriver.ChromeOptions()
prefs = {"profile.managed_default_content_settings.images":2}
chromeOptions.add_experimental_option("prefs",prefs)
chromeOptions.add_argument("--no-sandbox")
chromeOptions.add_extension('anticaptcha.crx')
chromeOptions.add_argument("--disable-setuid-sandbox")
driver = webdriver.Chrome(chrome_options=chromeOptions)
#driver = webdriver.Chrome(chrome_options=chromeOptions)
driver.get('chrome-extension://neodgnejhhhlcdoglifbmioajmagpeci/options.html')
driver.find_element_by_id('auto_submit_form').click()
driver.find_element_by_name('account_key').send_keys('5f6b75f0738a1bd09cf717392c9804a3')
driver.find_element_by_id('save').click()

driver.get('https://www.google.co.in/search?site=&source=hp&q=d')


for i in range(1505

        ,ws.max_row+1):
    print ws.cell(row=i, column=2).value
    try:
        driver.find_element_by_css_selector('.gsfi[title="Search"]').clear()
        driver.find_element_by_css_selector('.gsfi[title="Search"]').send_keys(ws.cell(row=i, column=2).value+'  '
                                                                               +ws.cell(row=i, column=3).value+'  '
                                                                               +ws.cell(row=i, column=4).value+'  '
                                                                               +str(ws.cell(row=i, column=5).value))
        driver.find_element_by_css_selector('.gsfi[title="Search"]').send_keys(Keys.RETURN)
        time.sleep(3)

        ws.cell(row=i, column=7).value = driver.find_element_by_css_selector('div.kno-ecr-pt.kno-fb-ctx._hdf > span:nth-child(1)').text
        ws.cell(row=i, column=9).value = driver.find_element_by_css_selector('span._Xbe._ZWk.kno-fv [data-dtype="d3ph"] span').text
        ws.cell(row=i, column=8).value = driver.find_element_by_css_selector('._b1m.kp-hc a.ab_button[role="button"]').get_attribute('href')
        wb.save('Manufacturing Companies_original.xlsx')
    except Exception,e:
        print e
        pass

